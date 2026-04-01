#!/usr/bin/env python3
"""Genera una struttura di cartelle numerata.

Esempi:
  python create_tree.py --base-dir . --root-name Cliente --children Contratti Fatture
  python create_tree.py --root-name Progetto --paths "Area tecnica/API" "Area tecnica/Test" "Vendite"
  python create_tree.py --root-name Cliente --excel struttura.xlsx
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, Iterable, List


Tree = Dict[str, "Tree"]


def sanitize_name(name: str) -> str:
    """Rimuove caratteri problematici per nomi cartelle."""
    cleaned = re.sub(r"[\\/:*?\"<>|]", "_", name.strip())
    cleaned = re.sub(r"\s+", " ", cleaned)
    if not cleaned:
        raise ValueError("Nome cartella vuoto o non valido.")
    return cleaned


def add_path(tree: Tree, parts: Iterable[str]) -> None:
    node = tree
    for part in parts:
        clean = sanitize_name(part)
        node = node.setdefault(clean, {})


def build_tree(children: List[str], paths: List[str], excel_rows: List[List[str]]) -> Tree:
    tree: Tree = {}

    for child in children:
        add_path(tree, [child])

    for raw_path in paths:
        parts = [p for p in raw_path.split("/") if p.strip()]
        if not parts:
            continue
        add_path(tree, parts)

    for row in excel_rows:
        parts = [p for p in row if p and str(p).strip()]
        if parts:
            add_path(tree, [str(p) for p in parts])

    return tree


def read_excel_rows(path: Path) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Per usare --excel installa openpyxl: pip install openpyxl"
        ) from exc

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if values:
            rows.append(values)
    return rows


def create_directories(base_dir: Path, root_name: str, tree: Tree, dry_run: bool = False) -> List[Path]:
    root_folder = base_dir / f"01_-{sanitize_name(root_name)}"
    created: List[Path] = [root_folder]

    def walk(node: Tree, parent_path: Path, prefix: str) -> None:
        for idx, (name, children) in enumerate(sorted(node.items()), start=1):
            code = f"{idx:02d}"
            folder_name = f"{prefix}{code}-{name}"
            current_path = parent_path / folder_name
            created.append(current_path)
            walk(children, current_path, f"{prefix}{code}_")

    walk(tree, root_folder, "01_")

    if not dry_run:
        for directory in created:
            directory.mkdir(parents=True, exist_ok=True)

    return created


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Crea una struttura di cartelle numerata (01_-, 01_01-, 01_01_01-, ...)."
    )
    parser.add_argument("--root-name", required=True, help="Nome della cartella radice.")
    parser.add_argument("--base-dir", default=".", help="Cartella base in cui creare l'albero.")
    parser.add_argument(
        "--children",
        nargs="*",
        default=[],
        help="Nomi cartelle di primo livello (es: --children Fatture Contratti).",
    )
    parser.add_argument(
        "--paths",
        nargs="*",
        default=[],
        help='Percorsi gerarchici separati da "/" (es: "Area/API" "Area/Test").',
    )
    parser.add_argument(
        "--excel",
        help="File .xlsx: ogni riga rappresenta un percorso gerarchico in colonne successive.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Mostra le cartelle senza crearle.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_rows: List[List[str]] = []

    if args.excel:
        excel_rows = read_excel_rows(Path(args.excel))

    tree = build_tree(args.children, args.paths, excel_rows)
    created = create_directories(Path(args.base_dir), args.root_name, tree, dry_run=args.dry_run)

    print("Cartelle generate:" if not args.dry_run else "Anteprima cartelle:")
    for folder in created:
        print(f" - {folder}")


if __name__ == "__main__":
    main()
